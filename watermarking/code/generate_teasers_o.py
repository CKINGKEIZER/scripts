"""
generate_teasers.py
-------------------
Reads buyers.xlsx:
  col A = recipient name
  col B = owner (editing) password  — blocks editing / copying
  col C = view password (optional)  — leave blank to keep the file openable without a password

Watermarks the source PDF and writes one encrypted PDF per recipient to the output folder.

Drop any PDF into the originals/ folder. Its filename becomes the prefix.
Example: "Project Sparkling - Teaser.pdf" -> "Project Sparkling - Teaser - Waterland.pdf"
"""

import os
import io
import sys
import urllib.request
import tarfile
import tempfile

# ── CONFIG ────────────────────────────────────────────────────────────────────
SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))   # watermarking/code/
PARENT_DIR       = os.path.dirname(SCRIPT_DIR)                   # watermarking/
EXCEL_PATH       = os.path.join(SCRIPT_DIR, "buyers.xlsx")
ORIGINALS_FOLDER = os.path.join(SCRIPT_DIR, "originals")
OUTPUT_FOLDER    = os.path.join(PARENT_DIR, "output")
FONT_PATH        = os.path.join(SCRIPT_DIR, "SourceSansPro-Regular.ttf")
# ─────────────────────────────────────────────────────────────────────────────


def check_dependencies():
    missing = []
    for pkg in ("pypdf", "reportlab", "openpyxl", "fontTools"):
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"Missing packages: {', '.join(missing)}")
        print(f"Run:  pip install {' '.join(missing)}")
        sys.exit(1)


def find_source_pdf(folder):
    """Return (full_path, filename_without_extension). Errors if 0 or 2+ PDFs found."""
    if not os.path.isdir(folder):
        print(f"ERROR: originals folder not found: {folder}")
        sys.exit(1)
    pdfs = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
    if len(pdfs) == 0:
        print(f"ERROR: No PDF found in {folder}")
        sys.exit(1)
    if len(pdfs) > 1:
        print(f"ERROR: Multiple PDFs found in {folder} — leave only one:")
        for f in pdfs:
            print(f"  {f}")
        sys.exit(1)
    name_without_ext = os.path.splitext(pdfs[0])[0]
    return os.path.join(folder, pdfs[0]), name_without_ext


def get_font(font_path):
    if os.path.isfile(font_path):
        return
    print("Font not found locally. Downloading Source Sans Pro...")
    url = "https://registry.npmjs.org/@fontsource/source-sans-pro/-/source-sans-pro-5.2.5.tgz"
    with tempfile.TemporaryDirectory() as tmpdir:
        tgz_path = os.path.join(tmpdir, "ssp.tgz")
        urllib.request.urlretrieve(url, tgz_path)
        with tarfile.open(tgz_path, "r:gz") as tar:
            member = next(
                m for m in tar.getmembers()
                if "source-sans-pro-latin-400-normal.woff" in m.name
            )
            tar.extract(member, tmpdir)
            woff_path = os.path.join(tmpdir, member.name)
        from fontTools.ttLib import TTFont as FTFont
        font = FTFont(woff_path)
        font.flavor = None
        font.save(font_path)
    print("Font ready.")


def register_font(font_path):
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont as RLFont
    pdfmetrics.registerFont(RLFont("SourceSansPro", font_path))


def read_excel(path):
    """
    Returns a list of (name, owner_password, view_password) tuples.
    view_password is "" when column C is empty or absent.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    buyers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name           = row[0] if len(row) > 0 else None
        owner_password = row[1] if len(row) > 1 else None
        view_password  = row[2] if len(row) > 2 else None

        if not name:
            continue
        owner_password = str(owner_password).strip() if owner_password else ""
        view_password = str(view_password).strip() if view_password else ""
        buyers.append((str(name).strip(), owner_password, view_password))
    wb.close()
    return buyers


def _max_chars_for_size(font_size):
    """
    Return the maximum number of characters allowed at a given font size.
    Pre-computed against A4 diagonal minus margins.
    """
    A4_USABLE = 931
    avg_char  = 0.48 * font_size
    return max(1, int(A4_USABLE / avg_char))


def _pick_font_size(buyer_name, chosen_size=None):
    """
    Return the font size to use for this name.
    If chosen_size is given, use it when the name fits, otherwise fall back
    through standard sizes. Floor is always 48.
    """
    STANDARD_SIZES = [72, 60, 48]

    if chosen_size is None:
        return 48 if len(buyer_name) > 26 else 72

    candidates = [chosen_size]
    for s in STANDARD_SIZES:
        if s < chosen_size and s not in candidates:
            candidates.append(s)
    if 48 not in candidates:
        candidates.append(48)

    for size in candidates:
        if len(buyer_name) <= _max_chars_for_size(size):
            return size
    return 48


def make_watermark(buyer_name, page_width, page_height, chosen_size=None):
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import Color
    from pypdf import PdfReader

    font_size = _pick_font_size(buyer_name, chosen_size)

    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(page_width, page_height))
    c.setFont("SourceSansPro", font_size)
    c.setFillColor(Color(0.5, 0.5, 0.5, alpha=0.15))
    c.saveState()
    c.translate(page_width / 2, page_height / 2)
    c.rotate(35)
    c.drawCentredString(0, 0, buyer_name)
    c.restoreState()
    c.save()
    packet.seek(0)
    return PdfReader(packet).pages[0]


def process(buyers, source_path, file_prefix, output_folder, chosen_size=None):
    from pypdf import PdfReader, PdfWriter
    from pypdf.constants import UserAccessPermissions

    PERMS = (
        UserAccessPermissions.PRINT |
        UserAccessPermissions.PRINT_TO_REPRESENTATION |
        UserAccessPermissions.EXTRACT_TEXT_AND_GRAPHICS
    )

    os.makedirs(output_folder, exist_ok=True)
    total = len(buyers)

    if chosen_size is not None:
        print(f"Watermark size: {chosen_size}pt  (auto-reduces for long names)")
        print()

    for i, (name, owner_password, view_password) in enumerate(buyers, 1):
        safe_name = name.replace("/", "-").replace("\\", "-").replace(":", "-")
        out_path  = os.path.join(output_folder, f"{file_prefix} - {safe_name}.pdf")

        reader = PdfReader(source_path)
        writer = PdfWriter()

        for page in reader.pages:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            page.merge_page(make_watermark(name, w, h, chosen_size))
            writer.add_page(page)

        if owner_password or view_password:
            writer.encrypt(
                user_password=view_password,   # "" = no password required to open
                owner_password=owner_password if owner_password else view_password,
                algorithm="AES-256",
                permissions_flag=PERMS,
            )

        labels = []
        if view_password:
            labels.append("view locked")
        if owner_password:
            labels.append("edit locked")
        view_label = " + ".join(labels) if labels else "watermark only"
        with open(out_path, "wb") as f:
            writer.write(f)

        print(f"  [{i}/{total}] {safe_name}  ({view_label})")

    print(f"\nDone. {total} files in: {output_folder}")


def main():
    check_dependencies()

    # Parse optional --font-size N argument
    chosen_size = None
    if "--font-size" in sys.argv:
        idx = sys.argv.index("--font-size")
        if idx + 1 < len(sys.argv):
            try:
                chosen_size = int(sys.argv[idx + 1])
            except ValueError:
                print(f"ERROR: --font-size must be an integer, got '{sys.argv[idx + 1]}'")
                sys.exit(1)

    if not os.path.isfile(EXCEL_PATH):
        print(f"ERROR: Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    source_path, file_prefix = find_source_pdf(ORIGINALS_FOLDER)

    print(f"Source PDF : {source_path}")
    print(f"Prefix     : {file_prefix}")
    print(f"Excel      : {EXCEL_PATH}")
    print(f"Output     : {OUTPUT_FOLDER}")
    print()

    get_font(FONT_PATH)
    register_font(FONT_PATH)

    buyers = read_excel(EXCEL_PATH)
    print(f"Recipients loaded: {len(buyers)}\n")

    process(buyers, source_path, file_prefix, OUTPUT_FOLDER, chosen_size)


if __name__ == "__main__":
    main()
