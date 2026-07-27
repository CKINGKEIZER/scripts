"""
create_folders.py
-----------------
Engine module: turn a pasted numbered outline into a nested folder tree.

Numbering determines depth:
  1          -> level 0
  1.1        -> level 1
  1.1.1      -> level 2   (and so on)

Used by:
  * launcher.py "Create Folders" tab
        parse_structure() for the live preview,
        create_folders() to build the tree + write folder_structure.xlsx
  * command line
        python create_folders.py structure.txt   (reads outline from a file)

Public functions:
  parse_structure(raw_text)                 -> [(depth, label), ...]
  build_folder_paths(entries, output_folder)-> [absolute path, ...]
  create_folders(paths, entries, output_folder, excel_dir, log) -> (created, skipped)
"""

import os
import re
import sys

# ── CONFIG ────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR    = os.path.dirname(SCRIPT_DIR)
OUTPUT_FOLDER = os.path.join(PARENT_DIR, "output")
# ─────────────────────────────────────────────────────────────────────────────

_LINE_RE = re.compile(r'^(\d+(?:\.\d+)*\.?)\s+(.+)')


def parse_structure(raw_text):
    """
    Parse a numbered outline into (depth, folder_name) pairs.
    Depth = number of dots in the numeric prefix.
      "1 Juridisch"       -> depth 0
      "1.1. Vennootschap" -> depth 1
      "1.1.1. Structuur"  -> depth 2
    Lines without a leading numeric prefix are skipped.
    Characters Windows forbids in folder names are replaced with '-'.
    """
    results = []
    for raw_line in raw_text.splitlines():
        # Collapse tabs (Excel multi-column paste) and extra spaces into one space.
        line = re.sub(r'[\t ]+', ' ', raw_line).strip()
        if not line:
            continue
        m = _LINE_RE.match(line)
        if not m:
            continue
        numeric_part = m.group(1).rstrip(".")
        # Rebuild label so no stray whitespace survives.
        label = f"{m.group(1).rstrip('.')}. {m.group(2).strip()}"
        label = re.sub(r'[<>:"/\\|?*]', '-', label).strip()
        # Windows forbids names ending with a period or space.
        label = label.rstrip('. ')
        depth = numeric_part.count(".")
        results.append((depth, label))
    return results


def build_folder_paths(entries, output_folder=OUTPUT_FOLDER):
    """Convert (depth, name) pairs to absolute folder paths under output_folder."""
    stack = {}
    paths = []
    for depth, name in entries:
        stack[depth] = name
        for k in [k for k in stack if k > depth]:
            del stack[k]
        parts = [stack[d] for d in sorted(stack)]
        paths.append(os.path.join(output_folder, *parts))
    return paths


def _save_structure_excel(entries, excel_dir, log):
    """Write a styled folder_structure.xlsx into excel_dir (for reference/editing)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folder Structure"
    ws.column_dimensions["A"].width = 120

    LEVEL_COLOURS = ["#0B2340", "#1A4A7A", "#2E6DA4", "#4A90C4"]
    LEVEL_SIZES   = [12, 11, 10, 9]

    for row_idx, (depth, label) in enumerate(entries, 1):
        cell = ws.cell(row=row_idx, column=1, value=label)
        hex_col   = LEVEL_COLOURS[min(depth, len(LEVEL_COLOURS) - 1)]
        font_size = LEVEL_SIZES[min(depth, len(LEVEL_SIZES) - 1)]
        cell.font = Font(name="Segoe UI", size=font_size, bold=(depth == 0), color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=hex_col.lstrip("#"))
        cell.alignment = Alignment(indent=depth * 2, vertical="center")
        ws.row_dimensions[row_idx].height = 18

    excel_path = os.path.join(excel_dir, "folder_structure.xlsx")
    wb.save(excel_path)
    return excel_path


def create_folders(paths, entries, output_folder=OUTPUT_FOLDER,
                   excel_dir=SCRIPT_DIR, log=None):
    """
    Create every folder in `paths`. Existing folders are left in place.
    Also writes folder_structure.xlsx into excel_dir. Returns (created, skipped).
    """
    log = log or print
    os.makedirs(output_folder, exist_ok=True)
    created = 0
    skipped = 0

    for path in paths:
        rel = os.path.relpath(path, os.path.dirname(output_folder))
        if os.path.isdir(path):
            log(f"  EXISTS   {rel}")
            skipped += 1
        else:
            os.makedirs(path, exist_ok=True)
            log(f"  CREATED  {rel}")
            created += 1

    try:
        _save_structure_excel(entries, excel_dir, log)
        log("\n  Structure saved  →  code/folder_structure.xlsx")
    except Exception as xe:
        log(f"\n  WARNING: could not save Excel  —  {xe}")

    log("")
    log(f"  Done.  {created} created,  {skipped} already existed.")
    return created, skipped


def _cli():
    if len(sys.argv) < 2:
        print("Usage: python create_folders.py <outline.txt>")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        raw = f.read()
    entries = parse_structure(raw)
    if not entries:
        print("No numbered outline found in the file.")
        sys.exit(1)
    paths = build_folder_paths(entries, OUTPUT_FOLDER)
    create_folders(paths, entries, OUTPUT_FOLDER)


if __name__ == "__main__":
    _cli()
