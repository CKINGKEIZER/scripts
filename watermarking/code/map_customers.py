"""
map_customers.py
----------------
Persistent customer-name mapping tool.

mapping_code.xlsx (lives alongside this script, never wiped):
  Sheet "Mapping" (sorted alphabetically by canonical_name):
    Col A = raw_name         every name variant ever seen
    Col B = canonical_name   the grouped/clean name
    Col C = customer_id      "Customer 1", "Customer 2", ...
    Col D = confidence        100 for exact, fuzzy score, "contain", or "new"
    Col E = source_file       which Excel the entry was first seen in

  Sheet "Output" (accumulates within a batch, wiped on New Batch):
    Col A = original value from the input file
    Col B = mapped result  ("Customer X" or pass-through)
    Col C = source file

Matching pipeline (per name):
  1. Strip order codes (OF-NNNNNN, standalone 4+ digit blocks)
  2. Exact match on raw_name                         -> reuse customer_id
  3. Fuzzy match on normalized canonical names        -> add to group
  4. Containment match (one name inside the other)    -> add to group
  5. No match                                         -> new customer number
"""

import os
import sys
import re

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
MAPPING_PATH = os.path.join(SCRIPT_DIR, "mapping_code.xlsx")


# ── DEPENDENCIES ──────────────────────────────────────────────────────────────

def check_dependencies(log=print):
    missing = []
    for pkg, pip_name in [("openpyxl", "openpyxl"), ("rapidfuzz", "rapidfuzz")]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pip_name)
    if missing:
        log(f"  Missing packages: {', '.join(missing)}")
        log(f"  Run: pip install {' '.join(missing)}")
        return False
    return True


# ── ORDER-CODE STRIPPING ─────────────────────────────────────────────────────

def strip_order_codes(name):
    """
    Remove order/reference codes. Keep company names, suffixes, descriptors.

    Handles:
      OF-240448           -> removed
      OF-240419-INF       -> removed
      240437              -> removed (standalone 4+ digit block)

    Does NOT remove:
      B.V.  LTD  renewal  new  TECS  etc.
    """
    s = str(name).strip()
    s = re.sub(r'\bOF-\d+(?:-[A-Za-z]+)*\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{4,}\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ── NOISE-TOKEN REMOVAL (fuzzy + containment matching only) ──────────────────

NOISE_TOKENS = {
    # Legal entity suffixes
    "b.v.", "bv", "b.v", "n.v.", "nv", "n.v", "ltd", "sa", "s.a.",
    "gmbh", "inc", "inc.", "ag", "se", "srl", "bvba", "vof", "cv",
    # Project descriptors
    "renewal", "new", "lease",
    # Project type codes
    "aec", "sdc", "tecs", "tcs", "hpc", "inf",
    # Location qualifiers (stripped for matching, kept in stored name)
    "belgium", "uk", "nederland", "netherlands", "europe", "europa",
    "global", "international",
    # Filler
    "short",
}


def normalize_for_matching(name):
    """
    Strip noise tokens so comparison focuses on the core company name.
    Only used for score calculation, never for storage or display.
    """
    tokens = name.lower().split()
    cleaned = [t for t in tokens if t.rstrip(".") not in NOISE_TOKENS and t not in NOISE_TOKENS]
    return " ".join(cleaned).strip()


# ── CONTAINMENT CHECK ────────────────────────────────────────────────────────

MIN_CONTAIN_LEN = 3   # minimum characters for the shorter name to count


def _contains_as_words(short, long):
    """True if every word in 'short' appears in 'long' (order-independent)."""
    s_words = set(short.split())
    l_words = set(long.split())
    return s_words.issubset(l_words)


def containment_match(norm_input, canon_to_cid):
    """
    Check if norm_input is fully contained in (or fully contains) any
    existing canonical name. Catches 'dana' matching 'dana belgium'.

    Returns (customer_id, canonical_name) or (None, None).
    """
    if len(norm_input) < MIN_CONTAIN_LEN:
        return None, None

    for canon_lower, (cid, canon_orig) in canon_to_cid.items():
        norm_canon = normalize_for_matching(canon_lower)
        if not norm_canon or len(norm_canon) < MIN_CONTAIN_LEN:
            continue

        # One must be a subset of the other's words
        if _contains_as_words(norm_input, norm_canon) or _contains_as_words(norm_canon, norm_input):
            return cid, canon_orig

    return None, None


# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────

def get_headers(excel_path):
    """Return list of header strings from row 1."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [
            str(h).strip() if h is not None else f"Column {i + 1}"
            for i, h in enumerate(row)
        ]
    wb.close()
    return headers


# ── LOAD / SAVE MAPPING ──────────────────────────────────────────────────────

def load_mapping(log=print):
    """
    Load persistent mapping from mapping_code.xlsx Sheet "Mapping".

    Returns
    -------
    entries         : list of dict
    next_num        : int
    existing_output : list of tuples (original, mapped, source)
    """
    import openpyxl

    if not os.path.isfile(MAPPING_PATH):
        log("  No existing mapping file found. Starting fresh.")
        return [], 1, []

    try:
        wb = openpyxl.load_workbook(MAPPING_PATH, read_only=True, data_only=True)
    except Exception as e:
        log(f"  ERROR reading mapping file: {e}")
        return [], 1, []

    # ── Mapping sheet ────────────────────────────────────────────
    entries = []
    max_num = 0

    if "Mapping" in wb.sheetnames:
        ws = wb["Mapping"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            raw    = str(row[0]).strip() if row[0] is not None else ""
            canon  = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            cid    = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            conf   = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            source = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

            if not raw:
                continue

            entries.append({
                "raw_name":       raw,
                "canonical_name": canon,
                "customer_id":    cid,
                "confidence":     conf,
                "source_file":    source,
            })

            m = re.match(r'Customer\s+(\d+)', cid, re.IGNORECASE)
            if m:
                max_num = max(max_num, int(m.group(1)))

    # ── Output sheet (for batch accumulation) ────────────────────
    existing_output = []
    if "Output" in wb.sheetnames:
        ws_out = wb["Output"]
        for i, row in enumerate(ws_out.iter_rows(values_only=True)):
            if i == 0:
                continue
            orig   = str(row[0]).strip() if row[0] is not None else ""
            mapped = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            src    = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            existing_output.append((orig, mapped, src))

    wb.close()
    log(f"  Loaded {len(entries)} mapping entries (max Customer {max_num}), "
        f"{len(existing_output)} existing output rows.")
    return entries, max_num + 1, existing_output


def save_mapping(entries, output_rows, log=print):
    """
    Write mapping_code.xlsx:
      Sheet "Mapping" — sorted alphabetically by canonical_name
      Sheet "Output"  — accumulated output rows for the current batch
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: Mapping ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Mapping"

    hdr_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(fill_type="solid", fgColor="0B2340")

    ws.append(["raw_name", "canonical_name", "customer_id", "confidence", "source_file"])
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A2"

    # Sort alphabetically by canonical_name, then raw_name within each group
    def _sort_key(e):
        return (e.get("canonical_name", "").lower(), e.get("raw_name", "").lower())

    sorted_entries = sorted(entries, key=_sort_key)

    ALT = ["EEF2F7", "FFFFFF"]
    prev_canon = None
    color_idx  = 0

    for entry in sorted_entries:
        if entry["canonical_name"].lower() != (prev_canon or "").lower():
            color_idx  = 1 - color_idx
            prev_canon = entry["canonical_name"]

        ws.append([
            entry["raw_name"],
            entry["canonical_name"],
            entry["customer_id"],
            entry["confidence"],
            entry.get("source_file", ""),
        ])

        fill = PatternFill(fill_type="solid", fgColor=ALT[color_idx])
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = Font(name="Segoe UI", size=9)

    # ── Sheet 2: Output ──────────────────────────────────────────────
    ws_out = wb.create_sheet("Output")
    ws_out.append(["original", "mapped", "source_file"])
    for cell in ws_out[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    ws_out.column_dimensions["A"].width = 50
    ws_out.column_dimensions["B"].width = 22
    ws_out.column_dimensions["C"].width = 30

    for orig, mapped, src in output_rows:
        ws_out.append([orig, mapped, src])

    wb.save(MAPPING_PATH)
    log(f"  Mapping saved  : {len(entries)} entries")
    log(f"  Output saved   : {len(output_rows)} rows on Sheet 'Output'")


# ── CLEAR OUTPUT (NEW BATCH) ─────────────────────────────────────────────────

def clear_output(log=print):
    """Wipe the Output sheet for a new batch. Mapping sheet is untouched."""
    import openpyxl

    if not os.path.isfile(MAPPING_PATH):
        log("  No mapping file exists yet. Nothing to clear.")
        return

    entries, _, _ = load_mapping(log=lambda _: None)
    save_mapping(entries, [], log)
    log("  Output sheet cleared for new batch.")


# ── MAIN PROCESS ─────────────────────────────────────────────────────────────

def process(input_path, col_index, threshold=85, log=print):
    """
    Run the mapping pipeline. Output appends to existing Output sheet
    (batch mode). Use clear_output() before a new batch.

    Returns
    -------
    (num_exact, num_fuzzy, num_contain, num_new, num_empty)
    """
    import openpyxl
    from rapidfuzz import fuzz

    if not check_dependencies(log):
        return (0, 0, 0, 0, 0)

    source_file = os.path.basename(input_path)

    # ── 1. Load existing mapping ─────────────────────────────────────
    entries, next_num, existing_output = load_mapping(log)

    # Build lookup indices (case-insensitive keys)
    raw_to_cid   = {}   # lowercase stripped name -> customer_id
    canon_to_cid = {}   # lowercase canonical     -> (customer_id, original_case_canon)

    for e in entries:
        raw_to_cid[e["raw_name"].lower()] = e["customer_id"]
        if e["canonical_name"]:
            key = e["canonical_name"].lower()
            canon_to_cid[key] = (e["customer_id"], e["canonical_name"])

    # ── 2. Read input column ─────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    except Exception as e:
        log(f"  ERROR opening input file: {e}")
        return (0, 0, 0, 0, 0)

    ws = wb.active
    input_values = []
    header_name  = f"Column {col_index + 1}"

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            if col_index < len(row) and row[col_index] is not None:
                header_name = str(row[col_index]).strip()
            continue
        val = row[col_index] if col_index < len(row) else None
        input_values.append(str(val).strip() if val is not None else "")
    wb.close()

    log(f"  Input file     : {source_file}")
    log(f"  Column         : '{header_name}' (index {col_index})")
    log(f"  Input rows     : {len(input_values)}")
    log(f"  Fuzzy threshold: {threshold}")
    log("")

    # ── 3. Match each name ───────────────────────────────────────────
    new_output   = []
    num_exact    = 0
    num_fuzzy    = 0
    num_contain  = 0
    num_new      = 0
    num_empty    = 0

    for raw_value in input_values:
        if not raw_value:
            new_output.append(("", "", source_file))
            num_empty += 1
            continue

        stripped = strip_order_codes(raw_value)

        if not stripped:
            new_output.append((raw_value, raw_value, source_file))
            log(f"  PASSTHROUGH  {raw_value}")
            num_empty += 1
            continue

        stripped_lower = stripped.lower()

        # 3a. Exact match on raw_name
        if stripped_lower in raw_to_cid:
            cid = raw_to_cid[stripped_lower]
            new_output.append((raw_value, cid, source_file))
            num_exact += 1
            continue

        # 3b. Fuzzy match on normalized names
        best_score = 0
        best_canon = None
        best_cid   = None
        norm_input = normalize_for_matching(stripped)

        if norm_input:
            for canon_lower, (cid, canon_orig) in canon_to_cid.items():
                norm_canon = normalize_for_matching(canon_lower)
                if not norm_canon:
                    continue
                score = fuzz.token_sort_ratio(norm_input, norm_canon)
                if score > best_score:
                    best_score = score
                    best_canon = canon_orig
                    best_cid   = cid

        if best_score >= threshold:
            entries.append({
                "raw_name":       stripped,
                "canonical_name": best_canon,
                "customer_id":    best_cid,
                "confidence":     str(int(best_score)),
                "source_file":    source_file,
            })
            raw_to_cid[stripped_lower] = best_cid
            new_output.append((raw_value, best_cid, source_file))
            log(f"  FUZZY {int(best_score)}%  '{stripped}' -> {best_cid} ({best_canon})")
            num_fuzzy += 1
            continue

        # 3c. Containment match (catches "DANA" inside "DANA Belgium")
        if norm_input:
            c_cid, c_canon = containment_match(norm_input, canon_to_cid)
            if c_cid is not None:
                entries.append({
                    "raw_name":       stripped,
                    "canonical_name": c_canon,
                    "customer_id":    c_cid,
                    "confidence":     "contain",
                    "source_file":    source_file,
                })
                raw_to_cid[stripped_lower] = c_cid
                new_output.append((raw_value, c_cid, source_file))
                log(f"  CONTAIN  '{stripped}' -> {c_cid} ({c_canon})")
                num_contain += 1
                continue

        # 3d. No match -> new customer
        new_cid = f"Customer {next_num}"
        entries.append({
            "raw_name":       stripped,
            "canonical_name": stripped,
            "customer_id":    new_cid,
            "confidence":     "new",
            "source_file":    source_file,
        })
        raw_to_cid[stripped_lower]   = new_cid
        canon_to_cid[stripped_lower] = (new_cid, stripped)
        new_output.append((raw_value, new_cid, source_file))
        log(f"  NEW  '{stripped}' -> {new_cid}")
        next_num += 1
        num_new  += 1

    log("")
    log(f"  Exact matches      : {num_exact}")
    log(f"  Fuzzy matches      : {num_fuzzy}")
    log(f"  Containment matches: {num_contain}")
    log(f"  New customers      : {num_new}")
    log(f"  Empty/passthru     : {num_empty}")
    log("")

    # ── 4. Save (append to output) ───────────────────────────────────
    all_output = existing_output + new_output
    save_mapping(entries, all_output, log)

    return (num_exact, num_fuzzy, num_contain, num_new, num_empty)
