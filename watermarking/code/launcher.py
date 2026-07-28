"""
launcher.py
-----------
GUI frontend for Kumulus Partners PDF tools.
Place this file in the same folder as generate_teasers.py and remove_passwords.py.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import subprocess
import sys
import os
import shutil
import threading

SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR       = os.path.dirname(SCRIPT_DIR)
ORIGINALS_FOLDER = os.path.join(SCRIPT_DIR, "originals")
OUTPUT_FOLDER    = os.path.join(PARENT_DIR, "output")   # watermarking/output — matches the engine scripts
BUYERS_EXCEL     = os.path.join(SCRIPT_DIR, "buyers.xlsx")
WORD_EXCEL       = os.path.join(SCRIPT_DIR, "word_placeholder_excel.xlsx")

# Convert-to-PDF output subfolder. Leading underscore keeps it sorted on top.
PDF_SUBFOLDER    = "_pdf"


# ── ENGINE LOADER ─────────────────────────────────────────────────────────────
# Each tool's logic lives in its own module next to this file (add_passwords,
# remove_passwords, create_folders, word_to_pdf, fill_template, generate_teasers).
# The GUI stays thin and just calls into them.

def _load_engine(module_name):
    """Import (and hot-reload) an engine module that sits next to launcher.py."""
    if SCRIPT_DIR not in sys.path:
        sys.path.insert(0, SCRIPT_DIR)
    import importlib
    mod = importlib.import_module(module_name)
    importlib.reload(mod)
    return mod


def _load_create_folders():
    return _load_engine("create_folders")


# ── OPEN FOLDER (cross-platform) ──────────────────────────────────────────────

def open_folder(path):
    """Open a folder in the OS file browser. Best-effort, never raises."""
    try:
        if not path:
            return
        os.makedirs(path, exist_ok=True)
        if sys.platform.startswith("win"):
            os.startfile(path)              # noqa: only exists on Windows
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


# ── SETTINGS (remember last-used options between launches) ─────────────────────

SETTINGS_PATH = os.path.join(SCRIPT_DIR, "settings.json")


def load_settings():
    try:
        import json
        with open(SETTINGS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_settings(data):
    try:
        import json
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


# ── DEPENDENCY CHECK ──────────────────────────────────────────────────────────

def check_and_install_packages():
    pkg_map = {
        "pypdf":        "pypdf",
        "reportlab":    "reportlab",
        "openpyxl":     "openpyxl",
        "fontTools":    "fonttools",
        "brotli":       "brotli",
        "cryptography": "cryptography",
        "tkinterdnd2":  "tkinterdnd2",
        "docx":         "python-docx",
        "win32com":     "pywin32",
    }
    missing = [v for k, v in pkg_map.items() if not _can_import(k)]
    if not missing:
        return
    if not messagebox.askyesno(
        "Missing packages",
        "The following packages are missing and need to be installed:\n\n"
        f"  {', '.join(missing)}\n\nInstall now?"
    ):
        sys.exit(0)
    try:
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet"] + missing,
            capture_output=True, text=True
        )
        if r.returncode != 0:
            messagebox.showerror("Install failed", r.stderr[-600:])
            sys.exit(1)

        # pywin32 needs a post-install step to register its DLLs (pythoncom,
        # pywintypes). Without it, 'import win32com' fails even after pip
        # reports success. Run it, then require a restart because those DLLs
        # will not load into the already-running process.
        if "pywin32" in missing:
            try:
                subprocess.run(
                    [sys.executable, "-m", "pywin32_postinstall", "-install"],
                    capture_output=True, text=True
                )
            except Exception:
                pass
            messagebox.showinfo(
                "Restart needed",
                "Packages installed, including pywin32.\n\n"
                "Please close this window and start the app again."
            )
            sys.exit(0)

        messagebox.showinfo("Done", "Packages installed. The app will now start.")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        sys.exit(1)

def _can_import(name):
    try:
        __import__(name)
        return True
    except ImportError:
        return False


# ── DESIGN TOKENS ─────────────────────────────────────────────────────────────

# Kumulus palette: deep navy on white, gold accent, generous white space
C = {
    # Backgrounds
    "bg":          "#F7F7F5",      # off-white page
    "panel":       "#FFFFFF",      # card surface
    "header_bg":   "#0B2340",      # deep navy header
    "nav_bg":      "#FFFFFF",      # nav bar
    "drop_bg":     "#F0F4F8",      # drop zone fill
    # Borders
    "border":      "#E2E2DC",      # subtle card border
    "border_mid":  "#C8D0D8",      # slightly stronger
    "drop_border": "#B8C8D8",      # drop zone outline
    "divider":     "#E8E8E4",      # nav separator
    # Text
    "text":        "#0B1E30",      # near-black body
    "sub":         "#7A8A98",      # secondary / placeholder
    "nav_active":  "#0B2340",      # active nav label
    "nav_idle":    "#7A8A98",      # idle nav label
    # Accent
    "gold":        "#C19A50",      # gold accent (from branding feel)
    "gold_dk":     "#A07C38",
    # Button
    "btn":         "#0B2340",      # primary button
    "btn_hover":   "#0D2E50",
    "btn_dis":     "#A0B0C0",
    # Status
    "ok":          "#1A5C36",
    "err":         "#8B1A1A",
    # Log
    "log_bg":      "#0D1B2A",
    "log_fg":      "#C8D8E8",
    "log_dim":     "#4A6A8A",
}
FF    = "Segoe UI"          # body
FF_H  = "Segoe UI Semibold" # headings / labels (fallback handled by weight)


# ── APP FACTORY ───────────────────────────────────────────────────────────────

def _build_app(base_class, has_dnd):

    class App(base_class):

        def __init__(self):
            super().__init__()
            self._has_dnd = has_dnd
            self._settings = load_settings()   # remembered options from last launch
            self.title("Kumulus Partners — PDF Tools")
            self.geometry("980x780")
            self.minsize(880, 700)
            self.configure(bg=C["bg"])
            self._setup_styles()
            self._build_header()
            self._build_nav()
            self._build_content()
            self._show_screen("watermark")

        # ── STYLES ─────────────────────────────────────────────────────────

        def _setup_styles(self):
            s = ttk.Style(self)
            s.theme_use("clam")
            s.configure(".",
                background=C["bg"], foreground=C["text"],
                font=(FF, 10), borderwidth=0, relief="flat"
            )
            s.configure("TFrame",  background=C["bg"])
            s.configure("TLabel",  background=C["bg"], foreground=C["text"])
            s.configure("TEntry",
                fieldbackground="white", foreground=C["text"],
                bordercolor=C["border_mid"], lightcolor=C["border_mid"],
                darkcolor=C["border_mid"], insertcolor=C["text"],
                padding=(8, 6)
            )
            # Primary button
            s.configure("Primary.TButton",
                background=C["btn"], foreground="white",
                font=(FF, 10, "bold"), padding=(26, 11), relief="flat",
                borderwidth=0
            )
            s.map("Primary.TButton",
                background=[("active", C["btn_hover"]), ("disabled", C["btn_dis"])],
                foreground=[("disabled", "#FFFFFF")]
            )
            # Ghost (text) button
            s.configure("Ghost.TButton",
                background=C["panel"], foreground=C["sub"],
                font=(FF, 9), padding=(12, 7), relief="flat", borderwidth=0
            )
            s.map("Ghost.TButton",
                background=[("active", C["drop_bg"])],
                foreground=[("active", C["btn"])]
            )
            # Nav buttons
            s.configure("NavOn.TButton",
                background=C["nav_bg"], foreground=C["nav_active"],
                font=(FF, 10, "bold"), padding=(20, 10), relief="flat", borderwidth=0
            )
            s.configure("NavOff.TButton",
                background=C["nav_bg"], foreground=C["nav_idle"],
                font=(FF, 10), padding=(20, 10), relief="flat", borderwidth=0
            )
            s.map("NavOff.TButton",
                foreground=[("active", C["nav_active"])]
            )
            # Progress bar
            s.configure("Gold.Horizontal.TProgressbar",
                troughcolor=C["border"], background=C["gold"],
                thickness=3, borderwidth=0
            )

        # ── HEADER ─────────────────────────────────────────────────────────

        def _build_header(self):
            hdr = tk.Frame(self, bg=C["header_bg"], height=62)
            hdr.pack(fill="x")
            hdr.pack_propagate(False)

            # Left: wordmark
            left = tk.Frame(hdr, bg=C["header_bg"])
            left.pack(side="left", padx=28, pady=0, fill="y")

            tk.Label(left,
                     text="KUMULUS PARTNERS",
                     bg=C["header_bg"], fg="#FFFFFF",
                     font=(FF, 14, "bold")
                     ).pack(side="left", anchor="s", pady=(0, 2))

            # Vertical separator
            tk.Frame(hdr, bg="#2A4A6A", width=1).pack(
                side="left", fill="y", pady=14, padx=18)

            tk.Label(hdr,
                     text="PDF Distribution Tools",
                     bg=C["header_bg"], fg="#7A9AB8",
                     font=(FF, 9)
                     ).pack(side="left", anchor="s", pady=(0, 3))

            # Right: version tag
            tk.Label(hdr,
                     text="Internal Tool",
                     bg=C["header_bg"], fg="#3A5A78",
                     font=(FF, 8)
                     ).pack(side="right", padx=28, anchor="s", pady=(0, 3))

        # ── NAV ────────────────────────────────────────────────────────────

        def _build_nav(self):
            nav = tk.Frame(self, bg=C["nav_bg"])
            nav.pack(fill="x")

            # Gold accent line at the very top of nav
            tk.Frame(self, bg=C["gold"], height=2).pack(fill="x")

            # Pack nav after the accent — need to reorder
            nav.pack_forget()
            tk.Frame(self, bg=C["gold"], height=2).pack_forget()

            # Rebuild in correct order: accent bar, then nav
            tk.Frame(self, bg=C["gold"], height=2).pack(fill="x")
            nav = tk.Frame(self, bg=C["nav_bg"])
            nav.pack(fill="x")

            self._nav_btns    = {}
            self._nav_markers = {}
            for key, label in [
                ("watermark",    "Watermark & Encrypt"),
                ("removepass",   "Remove Passwords"),
                ("folders",      "Create Folders"),
                ("filltemplate", "Fill Template"),
                ("addpass",      "Add Passwords"),
                ("wordtopdf",    "Convert to PDF"),
            ]:
                col = tk.Frame(nav, bg=C["nav_bg"])
                col.pack(side="left")
                btn = ttk.Button(col, text=label,
                                 command=lambda k=key: self._show_screen(k))
                btn.pack()
                # Gold underline marker
                marker = tk.Frame(col, bg=C["gold"], height=2)
                self._nav_btns[key]    = btn
                self._nav_markers[key] = marker

            tk.Frame(self, bg=C["divider"], height=1).pack(fill="x")

        def _show_screen(self, name):
            for s in self._screens.values():
                s.pack_forget()
            self._screens[name].pack(fill="both", expand=True)
            for k, b in self._nav_btns.items():
                if k == name:
                    b.configure(style="NavOn.TButton")
                    self._nav_markers[k].pack(fill="x")
                else:
                    b.configure(style="NavOff.TButton")
                    self._nav_markers[k].pack_forget()

        # ── CONTENT CONTAINER ──────────────────────────────────────────────

        def _build_content(self):
            outer = tk.Frame(self, bg=C["bg"])
            outer.pack(fill="both", expand=True, padx=24, pady=20)
            self._screens = {
                "watermark":    self._build_watermark_screen(outer),
                "removepass":   self._build_removepass_screen(outer),
                "folders":      self._build_folders_screen(outer),
                "filltemplate": self._build_filltemplate_screen(outer),
                "addpass":      self._build_addpass_screen(outer),
                "wordtopdf":    self._build_wordtopdf_screen(outer),
            }

        # ── SHARED WIDGET HELPERS ──────────────────────────────────────────

        def _card(self, parent, **kw):
            """Flat white card with a thin border."""
            return tk.Frame(parent, bg=C["panel"],
                            highlightbackground=C["border"],
                            highlightthickness=1, **kw)

        def _card_header(self, card, title, subtitle=None):
            """Section label inside a card."""
            hdr = tk.Frame(card, bg=C["panel"])
            hdr.pack(fill="x", padx=16, pady=(14, 0))
            tk.Label(hdr, text=title.upper(),
                     bg=C["panel"], fg=C["sub"],
                     font=(FF, 7, "bold")).pack(anchor="w")
            if subtitle:
                tk.Label(hdr, text=subtitle,
                         bg=C["panel"], fg=C["text"],
                         font=(FF, 9)).pack(anchor="w", pady=(2, 0))
            # Thin rule under header
            tk.Frame(card, bg=C["divider"], height=1).pack(
                fill="x", padx=16, pady=(8, 0))

        def _dropzone(self, parent, icon, text_var, on_click):
            outer = tk.Frame(parent, bg=C["panel"])
            outer.pack(fill="both", expand=True, padx=16, pady=12)

            zone = tk.Frame(outer, bg=C["drop_bg"],
                            highlightbackground=C["drop_border"],
                            highlightthickness=1, cursor="hand2")
            zone.pack(fill="both", expand=True)

            inner = tk.Frame(zone, bg=C["drop_bg"])
            inner.place(relx=0.5, rely=0.5, anchor="center")

            tk.Label(inner, text=icon, bg=C["drop_bg"],
                     font=(FF, 22), fg=C["sub"]).pack()
            lbl = tk.Label(inner, textvariable=text_var,
                           bg=C["drop_bg"], fg=C["sub"],
                           font=(FF, 9), justify="center", wraplength=220)
            lbl.pack(pady=(4, 0))

            for w in [zone, inner, lbl]:
                w.bind("<Button-1>", on_click)
            return zone, lbl

        def _register_drop(self, widget, callback):
            if not self._has_dnd:
                return
            try:
                from tkinterdnd2 import DND_FILES
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", callback)
            except Exception:
                pass

        def _label(self, parent, text, size=9, bold=False, color=None):
            return tk.Label(parent,
                            text=text,
                            bg=C["panel"],
                            fg=color or C["sub"],
                            font=(FF, size, "bold" if bold else "normal"))

        def _log_widget(self, parent):
            log = scrolledtext.ScrolledText(
                parent, height=8,
                font=("Consolas", 8),
                bg=C["log_bg"], fg=C["log_fg"],
                insertbackground=C["log_fg"],
                relief="flat", state="disabled",
                borderwidth=0, highlightthickness=0,
                selectbackground="#1A3A5A",
                selectforeground=C["log_fg"],
            )
            log.pack(fill="both", expand=True, padx=16, pady=(8, 14))
            return log

        def _log_write(self, log, msg):
            log.configure(state="normal")
            log.insert("end", msg + "\n")
            log.see("end")
            log.configure(state="disabled")

        def _password_row(self, parent, pw_var, show_var):
            """Reusable password entry + show toggle."""
            row = tk.Frame(parent, bg=C["panel"])
            row.pack(fill="x")
            entry = ttk.Entry(row, textvariable=pw_var,
                              font=(FF, 10), show="●")
            entry.pack(side="left", fill="x", expand=True)

            def toggle():
                show_var.set(not show_var.get())
                entry.configure(show="" if show_var.get() else "●")

            ttk.Button(row, text="Show", style="Ghost.TButton",
                       width=5, command=toggle).pack(side="left", padx=(6, 0))
            return entry

        def _open_output_button(self, parent, path_getter):
            """A ghost button that opens the tool's output folder in Explorer."""
            return ttk.Button(
                parent, text="📂  Open output folder", style="Ghost.TButton",
                command=lambda: open_folder(path_getter()))

        def _persist_settings(self):
            """Remember the options that are annoying to re-set each launch."""
            data = dict(self._settings)
            for key, getter in (
                ("wm_size",        lambda: self._wm_size_var.get()),
                ("wp_strip_index", lambda: bool(self._wp_strip_index.get())),
                ("wp_recurse",     lambda: bool(self._wp_recurse.get())),
                ("wp_skip_pdfs",   lambda: bool(self._wp_skip_pdfs.get())),
                ("ft_to_pdf",      lambda: bool(self._ft_pdf_var.get())),
            ):
                try:
                    data[key] = getter()
                except Exception:
                    pass
            self._settings = data
            save_settings(data)

        def _check_output_folder(self):
            if not os.path.isdir(OUTPUT_FOLDER):
                return True
            existing = [f for f in os.listdir(OUTPUT_FOLDER)
                        if os.path.isfile(os.path.join(OUTPUT_FOLDER, f))]
            if not existing:
                return True
            ok = messagebox.askyesno(
                "Output folder is not empty",
                f"The output folder contains {len(existing)} file(s) from a previous run.\n\n"
                "Have you already moved them to a safe location?\n\n"
                "Yes  →  clear the output folder and continue\n"
                "No   →  cancel",
                icon="warning"
            )
            if not ok:
                return False
            for f in existing:
                try:
                    os.remove(os.path.join(OUTPUT_FOLDER, f))
                except Exception:
                    pass
            return True

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 1 — WATERMARK & ENCRYPT
        # ═══════════════════════════════════════════════════════════════════

        def _build_watermark_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            # ── Row 1: two cards side by side ──
            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="x", pady=(0, 14))

            # Card A: Source PDF
            card_pdf = self._card(row1)
            card_pdf.pack(side="left", fill="both", expand=True, padx=(0, 12))
            self._card_header(card_pdf, "Source PDF",
                              "The document to be watermarked and encrypted")

            self._wm_pdf_path = None
            self._wm_pdf_var  = tk.StringVar(value="Drop PDF here  ·  or click to browse")
            wm_zone, self._wm_zone_lbl = self._dropzone(
                card_pdf, "📄", self._wm_pdf_var, self._wm_browse_pdf
            )
            self._register_drop(wm_zone, self._wm_on_drop)

            # Card B: Recipients + Editing password + Viewing password (stacked)
            right_col = tk.Frame(row1, bg=C["bg"])
            right_col.pack(side="left", fill="both", expand=True)

            # Recipients
            card_names = self._card(right_col)
            card_names.pack(fill="both", expand=True, pady=(0, 12))
            self._card_header(card_names, "Recipients",
                              "Paste one name per line (direct from Excel column)")
            self._wm_names = scrolledtext.ScrolledText(
                card_names, height=5, font=(FF, 10),
                bg="white", fg=C["text"], relief="flat",
                highlightbackground=C["border"], highlightthickness=0,
                insertbackground=C["text"], wrap="none",
                borderwidth=0
            )
            self._wm_names.pack(fill="both", expand=True, padx=16, pady=(10, 14))

            # Editing password
            card_pw = self._card(right_col)
            card_pw.pack(fill="x", pady=(0, 12))

            # Header row: label on left, toggle on right
            edit_hdr_row = tk.Frame(card_pw, bg=C["panel"])
            edit_hdr_row.pack(fill="x", padx=16, pady=(14, 0))
            tk.Label(edit_hdr_row, text="EDITING PASSWORD",
                     bg=C["panel"], fg=C["sub"],
                     font=(FF, 7, "bold")).pack(side="left", anchor="w")
            self._wm_edit_enabled = tk.BooleanVar(value=False)
            tk.Checkbutton(
                edit_hdr_row,
                text="Enable",
                variable=self._wm_edit_enabled,
                bg=C["panel"], fg=C["sub"],
                activebackground=C["panel"],
                activeforeground=C["text"],
                selectcolor=C["panel"],
                font=(FF, 8),
                cursor="hand2",
                command=self._wm_toggle_edit_pw,
            ).pack(side="right", anchor="e")
            tk.Frame(card_pw, bg=C["divider"], height=1).pack(
                fill="x", padx=16, pady=(8, 0))

            # Subtitle always visible
            tk.Label(card_pw,
                     text="Blocks editing and copying — file opens without this password",
                     bg=C["panel"], fg=C["text"], font=(FF, 9)
                     ).pack(anchor="w", padx=16, pady=(8, 0))

            # Disabled state label (shown when toggle is off)
            self._wm_edit_disabled_lbl = tk.Label(
                card_pw,
                text="Currently disabled — no editing restriction applied",
                bg=C["panel"], fg=C["sub"], font=(FF, 8, "italic")
            )
            self._wm_edit_disabled_lbl.pack(anchor="w", padx=16, pady=(4, 14))

            # Password entry frame (hidden until toggle is on)
            self._wm_edit_pw_frame = tk.Frame(card_pw, bg=C["panel"])
            edit_pw_inner = tk.Frame(self._wm_edit_pw_frame, bg=C["panel"])
            edit_pw_inner.pack(fill="x", padx=16, pady=(6, 14))
            self._wm_pw_var  = tk.StringVar()
            self._wm_pw_show = tk.BooleanVar(value=False)
            self._password_row(edit_pw_inner, self._wm_pw_var, self._wm_pw_show)

            # Viewing password
            card_view = self._card(right_col)
            card_view.pack(fill="x")

            # Header row: label on left, toggle on right
            view_hdr_row = tk.Frame(card_view, bg=C["panel"])
            view_hdr_row.pack(fill="x", padx=16, pady=(14, 0))
            tk.Label(view_hdr_row, text="VIEWING PASSWORD",
                     bg=C["panel"], fg=C["sub"],
                     font=(FF, 7, "bold")).pack(side="left", anchor="w")
            self._wm_view_enabled = tk.BooleanVar(value=False)
            tk.Checkbutton(
                view_hdr_row,
                text="Enable",
                variable=self._wm_view_enabled,
                bg=C["panel"], fg=C["sub"],
                activebackground=C["panel"],
                activeforeground=C["text"],
                selectcolor=C["panel"],
                font=(FF, 8),
                cursor="hand2",
                command=self._wm_toggle_view_pw,
            ).pack(side="right", anchor="e")
            tk.Frame(card_view, bg=C["divider"], height=1).pack(
                fill="x", padx=16, pady=(8, 0))

            # Subtitle always visible
            tk.Label(card_view,
                     text="Recipients must enter this password to open the file",
                     bg=C["panel"], fg=C["text"], font=(FF, 9)
                     ).pack(anchor="w", padx=16, pady=(8, 0))

            # Disabled state label (shown when toggle is off)
            self._wm_view_disabled_lbl = tk.Label(
                card_view,
                text="Currently disabled — file opens without a password",
                bg=C["panel"], fg=C["sub"], font=(FF, 8, "italic")
            )
            self._wm_view_disabled_lbl.pack(anchor="w", padx=16, pady=(4, 14))

            # Password entry frame (hidden until toggle is on)
            self._wm_view_pw_frame = tk.Frame(card_view, bg=C["panel"])
            view_pw_inner = tk.Frame(self._wm_view_pw_frame, bg=C["panel"])
            view_pw_inner.pack(fill="x", padx=16, pady=(6, 14))
            self._wm_view_pw_var  = tk.StringVar()
            self._wm_view_pw_show = tk.BooleanVar(value=False)
            self._password_row(view_pw_inner, self._wm_view_pw_var, self._wm_view_pw_show)

            # ── Settings row: watermark size ──
            settings_row = tk.Frame(frame, bg=C["bg"])
            settings_row.pack(fill="x", pady=(0, 6))

            tk.Label(settings_row, text="Watermark size (pt):",
                     bg=C["bg"], fg=C["text"], font=(FF, 9)
                     ).pack(side="left")

            self._wm_size_var = tk.StringVar(
                value=str(self._settings.get("wm_size", "72")))
            size_combo = ttk.Combobox(
                settings_row, textvariable=self._wm_size_var,
                values=["48", "60", "72", "84", "92", "108"],
                width=5, font=(FF, 10), state="normal"
            )
            size_combo.pack(side="left", padx=(6, 8))

            tk.Label(settings_row,
                     text="Auto-reduces for long names",
                     bg=C["bg"], fg=C["sub"], font=(FF, 8)
                     ).pack(side="left")

            # ── Row 2: action bar ──
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(14, 14))

            self._wm_run_btn = ttk.Button(row2, text="Generate PDFs",
                                           style="Primary.TButton",
                                           command=self._wm_run)
            self._wm_run_btn.pack(side="left")

            self._wm_bar = ttk.Progressbar(
                row2, mode="indeterminate", length=160,
                style="Gold.Horizontal.TProgressbar"
            )

            self._wm_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._wm_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(
                row2, lambda: OUTPUT_FOLDER).pack(side="right")

            # ── Row 3: log card ──
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._wm_log = self._log_widget(card_log)

            return frame

        def _wm_toggle_view_pw(self):
            if self._wm_view_enabled.get():
                self._wm_view_disabled_lbl.pack_forget()
                self._wm_view_pw_frame.pack(fill="x")
            else:
                self._wm_view_pw_frame.pack_forget()
                self._wm_view_pw_var.set("")
                self._wm_view_disabled_lbl.pack(anchor="w", padx=16, pady=(4, 14))

        def _wm_toggle_edit_pw(self):
            if self._wm_edit_enabled.get():
                self._wm_edit_disabled_lbl.pack_forget()
                self._wm_edit_pw_frame.pack(fill="x")
            else:
                self._wm_edit_pw_frame.pack_forget()
                self._wm_pw_var.set("")
                self._wm_edit_disabled_lbl.pack(anchor="w", padx=16, pady=(4, 14))

        def _wm_on_drop(self, event):
            path = event.data.strip().strip("{}")
            if path.lower().endswith(".pdf"):
                self._wm_set_pdf(path)
            else:
                messagebox.showwarning("Invalid file", "Please drop a PDF file.")

        def _wm_browse_pdf(self, event=None):
            path = filedialog.askopenfilename(
                title="Select source PDF",
                filetypes=[("PDF files", "*.pdf")]
            )
            if path:
                self._wm_set_pdf(path)

        def _wm_set_pdf(self, path):
            self._wm_pdf_path = path
            self._wm_pdf_var.set(f"✓  {os.path.basename(path)}")
            self._wm_zone_lbl.configure(fg=C["ok"])

        def _wm_run(self):
            if not self._wm_pdf_path:
                messagebox.showwarning("Missing PDF", "Select a source PDF first.")
                return
            names = [n.strip() for n in
                     self._wm_names.get("1.0", "end").strip().splitlines()
                     if n.strip()]
            if not names:
                messagebox.showwarning("Missing names", "Paste at least one recipient name.")
                return
            pw = ""
            if self._wm_edit_enabled.get():
                pw = self._wm_pw_var.get().strip()
                if not pw:
                    messagebox.showwarning("Missing password", "Enter an editing password or disable the option.")
                    return

            view_pw = ""
            if self._wm_view_enabled.get():
                view_pw = self._wm_view_pw_var.get().strip()
                if not view_pw:
                    messagebox.showwarning(
                        "Missing viewing password",
                        "Enter a viewing password or disable the option."
                    )
                    return
                if pw and view_pw == pw:
                    messagebox.showwarning(
                        "Password conflict",
                        "The viewing password must be different from the editing password."
                    )
                    return

            if not self._check_output_folder():
                return

            # Validate watermark size
            font_size_str = self._wm_size_var.get().strip()
            try:
                font_size = int(font_size_str)
                if font_size < 12 or font_size > 200:
                    raise ValueError
            except ValueError:
                messagebox.showwarning(
                    "Invalid size",
                    "Watermark size must be a whole number between 12 and 200."
                )
                return

            self._wm_run_btn.configure(state="disabled")
            self._wm_bar.pack(side="left", padx=(14, 0))
            self._wm_bar.start(8)
            self._wm_status.set("Running…")
            self._wm_log.configure(state="normal")
            self._wm_log.delete("1.0", "end")
            self._wm_log.configure(state="disabled")

            self._persist_settings()
            threading.Thread(
                target=self._wm_execute, args=(names, pw, view_pw, font_size), daemon=True
            ).start()

        def _wm_execute(self, names, pw, view_pw="", font_size=72):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._wm_log, m))

            try:
                import openpyxl

                os.makedirs(ORIGINALS_FOLDER, exist_ok=True)
                for f in os.listdir(ORIGINALS_FOLDER):
                    if f.lower().endswith(".pdf"):
                        os.remove(os.path.join(ORIGINALS_FOLDER, f))
                dest = os.path.join(ORIGINALS_FOLDER, os.path.basename(self._wm_pdf_path))
                shutil.copy2(self._wm_pdf_path, dest)
                log(f"  PDF copied  →  originals/{os.path.basename(dest)}")

                wb = openpyxl.Workbook()
                ws = wb.active
                ws["A1"] = "Name"
                ws["B1"] = "Password Editing"
                ws["C1"] = "Password Viewing"
                for i, name in enumerate(names, 2):
                    ws.cell(row=i, column=1, value=name)
                    ws.cell(row=i, column=2, value=pw)
                    ws.cell(row=i, column=3, value=view_pw)  # "" = no open password
                wb.save(BUYERS_EXCEL)

                labels = []
                if view_pw:
                    labels.append("view locked")
                if pw:
                    labels.append("edit locked")
                view_note = " + ".join(labels) if labels else "watermark only"
                log(f"  buyers.xlsx written  ({len(names)} recipients, {view_note})")
                log("")

                log("  Running generate_teasers.py …")
                log("  " + "─" * 44)
                cmd = [sys.executable, os.path.join(SCRIPT_DIR, "generate_teasers.py"),
                       "--font-size", str(font_size)]
                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, cwd=SCRIPT_DIR
                )
                for line in proc.stdout:
                    log("  " + line.rstrip())
                proc.wait()

                if proc.returncode == 0:
                    done_note = " (viewing password set)" if view_pw else ""
                    self.after(0, lambda: self._wm_status.set(
                        f"✓  {len(names)} PDF(s) generated{done_note}"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Done",
                        f"{len(names)} PDF(s) generated.\n\nOutput folder:  output/"
                    ))
                else:
                    self.after(0, lambda: self._wm_status.set("Error — see log"))

            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._wm_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._wm_run_btn.configure(state="normal"))
                self.after(0, lambda: self._wm_bar.stop())
                self.after(0, lambda: self._wm_bar.pack_forget())

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 2 — REMOVE PASSWORDS
        # ═══════════════════════════════════════════════════════════════════

        def _build_removepass_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="x", pady=(0, 14))

            # Card: files — drop zone with explicit height so it always renders
            card_files = self._card(row1)
            card_files.pack(side="left", fill="both", expand=True, padx=(0, 12))
            self._card_header(card_files, "Encrypted PDFs",
                              "Select individual files or an entire folder")

            drop_wrapper = tk.Frame(card_files, bg=C["panel"], height=160)
            drop_wrapper.pack(fill="x", padx=16, pady=(10, 0))
            drop_wrapper.pack_propagate(False)

            self._rp_files = []
            self._rp_files_var = tk.StringVar(value="Drop files here  ·  or use Browse below")
            rp_zone, self._rp_zone_lbl = self._dropzone(
                drop_wrapper, "🔒", self._rp_files_var, self._rp_browse_files
            )
            self._register_drop(rp_zone, self._rp_on_drop)

            browse_row = tk.Frame(card_files, bg=C["panel"])
            browse_row.pack(fill="x", padx=16, pady=(8, 14))
            ttk.Button(browse_row, text="Browse files",
                       style="Ghost.TButton",
                       command=self._rp_browse_files).pack(side="left", padx=(0, 6))
            ttk.Button(browse_row, text="Browse folder",
                       style="Ghost.TButton",
                       command=self._rp_browse_folder).pack(side="left")

            # Card: password — fixed width so it never gets squeezed
            card_pw = self._card(row1, width=280)
            card_pw.pack(side="left", fill="y")
            card_pw.pack_propagate(False)
            self._card_header(card_pw, "Owner Password",
                              "The password used to encrypt these PDFs")
            pw_inner = tk.Frame(card_pw, bg=C["panel"])
            pw_inner.pack(fill="x", padx=16, pady=(10, 14))
            self._rp_pw_var  = tk.StringVar()
            self._rp_pw_show = tk.BooleanVar(value=False)
            self._password_row(pw_inner, self._rp_pw_var, self._rp_pw_show)

            # Action bar
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(0, 14))

            self._rp_run_btn = ttk.Button(row2, text="Remove Passwords",
                                           style="Primary.TButton",
                                           command=self._rp_run)
            self._rp_run_btn.pack(side="left")

            self._rp_bar = ttk.Progressbar(
                row2, mode="indeterminate", length=160,
                style="Gold.Horizontal.TProgressbar"
            )

            self._rp_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._rp_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(row2, lambda: OUTPUT_FOLDER).pack(side="right")

            # Log
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._rp_log = self._log_widget(card_log)

            return frame

        def _rp_on_drop(self, event):
            import re
            parts = re.findall(r'\{([^}]+)\}|(\S+)', event.data.strip())
            flat  = [p[0] or p[1] for p in parts]
            pdfs  = [p for p in flat if p.lower().endswith(".pdf")]
            if pdfs:
                self._rp_set_files(pdfs)
            else:
                messagebox.showwarning("No PDFs", "No PDF files detected.")

        def _rp_browse_files(self, event=None):
            paths = filedialog.askopenfilenames(
                title="Select encrypted PDFs",
                filetypes=[("PDF files", "*.pdf")]
            )
            if paths:
                self._rp_set_files(list(paths))

        def _rp_browse_folder(self):
            folder = filedialog.askdirectory(title="Select folder with encrypted PDFs")
            if not folder:
                return
            pdfs = [os.path.join(folder, f)
                    for f in os.listdir(folder) if f.lower().endswith(".pdf")]
            if not pdfs:
                messagebox.showwarning("No PDFs", "No PDF files found in that folder.")
                return
            self._rp_set_files(pdfs)

        def _rp_set_files(self, paths):
            self._rp_files = paths
            self._rp_files_var.set(f"✓  {len(paths)} file(s) selected")
            self._rp_zone_lbl.configure(fg=C["ok"])

        def _rp_run(self):
            if not self._rp_files:
                messagebox.showwarning("No files", "Select at least one PDF.")
                return
            pw = self._rp_pw_var.get().strip()
            if not pw:
                messagebox.showwarning("Missing password", "Enter the owner password.")
                return
            if not self._check_output_folder():
                return

            self._rp_run_btn.configure(state="disabled")
            self._rp_bar.pack(side="left", padx=(14, 0))
            self._rp_bar.start(8)
            self._rp_status.set("Running…")
            self._rp_log.configure(state="normal")
            self._rp_log.delete("1.0", "end")
            self._rp_log.configure(state="disabled")

            threading.Thread(
                target=self._rp_execute,
                args=(list(self._rp_files), pw),
                daemon=True
            ).start()

        def _rp_execute(self, files, pw):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._rp_log, m))

            try:
                if SCRIPT_DIR not in sys.path:
                    sys.path.insert(0, SCRIPT_DIR)
                import remove_passwords
                import importlib
                importlib.reload(remove_passwords)

                ok, total = remove_passwords.remove_passwords(
                    files, pw, OUTPUT_FOLDER, log=log)

                if ok == total:
                    self.after(0, lambda: self._rp_status.set(f"✓  {ok} file(s) done"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Done", f"{ok} file(s) decrypted.\n\nOutput folder:  output/"
                    ))
                else:
                    self.after(0, lambda: self._rp_status.set(
                        f"{ok}/{total} done — check log"))

            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._rp_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._rp_run_btn.configure(state="normal"))
                self.after(0, lambda: self._rp_bar.stop())
                self.after(0, lambda: self._rp_bar.pack_forget())

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 3 — CREATE FOLDERS
        # ═══════════════════════════════════════════════════════════════════

        # ── Parsing helpers ────────────────────────────────────────────────

        @staticmethod
        def _cf_parse(raw_text):
            """Parse a numbered outline into (depth, name) pairs (via engine)."""
            return _load_create_folders().parse_structure(raw_text)

        @staticmethod
        def _cf_build_paths(entries):
            """Convert (depth, name) pairs to absolute folder paths under OUTPUT_FOLDER."""
            return _load_create_folders().build_folder_paths(entries, OUTPUT_FOLDER)

        @staticmethod
        def _cf_preview(entries):
            lines = []
            for depth, name in entries:
                lines.append("    " * depth + name)
            return "\n".join(lines)

        # ── Screen builder ─────────────────────────────────────────────────

        def _build_folders_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            # ── Row 1: two panes side by side ──
            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="both", expand=True, pady=(0, 14))

            # Left: paste input
            card_in = self._card(row1)
            card_in.pack(side="left", fill="both", expand=True, padx=(0, 12))
            self._card_header(card_in, "Folder Structure",
                              "Paste your numbered outline — one item per line, no blank lines")

            self._cf_input = scrolledtext.ScrolledText(
                card_in, font=("Consolas", 9),
                bg="white", fg=C["text"], relief="flat",
                highlightbackground=C["border"], highlightthickness=0,
                insertbackground=C["text"], wrap="none", borderwidth=0
            )
            self._cf_input.pack(fill="both", expand=True, padx=16, pady=(10, 14))
            self._cf_input.bind("<KeyRelease>",  self._cf_on_change)
            self._cf_input.bind("<Control-v>",   self._cf_on_paste)
            self._cf_input.bind("<Control-V>",   self._cf_on_paste)

            # Right: live preview
            card_prev = self._card(row1)
            card_prev.pack(side="left", fill="both", expand=True)
            self._card_header(card_prev, "Preview",
                              "How the folders will nest inside output/")

            self._cf_preview_widget = scrolledtext.ScrolledText(
                card_prev, font=("Consolas", 9),
                bg=C["drop_bg"], fg=C["text"], relief="flat",
                highlightbackground=C["border"], highlightthickness=0,
                insertbackground=C["text"], wrap="none",
                borderwidth=0, state="disabled"
            )
            self._cf_preview_widget.pack(fill="both", expand=True, padx=16, pady=(10, 14))

            # ── Row 2: action bar ──
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(0, 14))

            self._cf_run_btn = ttk.Button(row2, text="Create Folders",
                                          style="Primary.TButton",
                                          command=self._cf_run)
            self._cf_run_btn.pack(side="left")

            self._cf_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._cf_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(row2, lambda: OUTPUT_FOLDER).pack(side="right")

            # ── Row 3: log ──
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._cf_log = self._log_widget(card_log)

            return frame

        def _cf_on_change(self, _event=None):
            self.after(50, self._cf_refresh_preview)

        def _cf_on_paste(self, _event=None):
            # Let tkinter complete the native paste first, then refresh.
            # Do NOT return "break" — that would block the paste itself.
            self.after(100, self._cf_refresh_preview)

        def _cf_refresh_preview(self):
            raw     = self._cf_input.get("1.0", "end")
            entries = self._cf_parse(raw)

            MAX_LABEL = 80   # chars; keeps total path safely under 260

            self._cf_preview_widget.configure(state="normal")
            self._cf_preview_widget.delete("1.0", "end")
            self._cf_preview_widget.tag_configure("ok",  foreground=C["text"])
            self._cf_preview_widget.tag_configure("bad", foreground="#C0392B")

            has_errors = False
            for depth, label in entries:
                indent = "    " * depth
                line   = indent + label + "\n"
                if len(label) > MAX_LABEL:
                    self._cf_preview_widget.insert("end", line, "bad")
                    has_errors = True
                else:
                    self._cf_preview_widget.insert("end", line, "ok")

            if has_errors:
                self._cf_preview_widget.insert(
                    "end",
                    "\n  ⚠  Red lines exceed 80 characters.\n"
                    "  Shorten them in your source before creating folders.\n",
                    "bad"
                )
                self._cf_run_btn.configure(state="disabled")
            else:
                self._cf_run_btn.configure(state="normal")

            self._cf_preview_widget.configure(state="disabled")

        def _cf_run(self):
            raw     = self._cf_input.get("1.0", "end")
            entries = self._cf_parse(raw)
            if not entries:
                messagebox.showwarning("No input",
                                       "Paste a numbered folder structure first.")
                return
            paths = self._cf_build_paths(entries)

            self._cf_run_btn.configure(state="disabled")
            self._cf_status.set("Creating…")
            self._cf_log.configure(state="normal")
            self._cf_log.delete("1.0", "end")
            self._cf_log.configure(state="disabled")

            threading.Thread(
                target=self._cf_execute, args=(paths, entries), daemon=True
            ).start()

        def _cf_execute(self, paths, entries):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._cf_log, m))

            try:
                cf = _load_create_folders()
                created, skipped = cf.create_folders(
                    paths, entries,
                    output_folder=OUTPUT_FOLDER,
                    excel_dir=SCRIPT_DIR,
                    log=log,
                )
                self.after(0, lambda: self._cf_status.set(
                    f"✓  {created} folder(s) created"))
                if created:
                    self.after(0, lambda: messagebox.showinfo(
                        "Done",
                        f"{created} folder(s) created.\n\nLocation:  output/\nStructure saved:  code/folder_structure.xlsx"
                    ))

            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._cf_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._cf_run_btn.configure(state="normal"))

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 4 — FILL TEMPLATE
        # ═══════════════════════════════════════════════════════════════════

        def _build_filltemplate_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            # ── Row 1: two cards side by side ──
            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="x", pady=(0, 14))

            # Card A: template picker + PDF toggle
            card_tpl = self._card(row1)
            card_tpl.pack(side="left", fill="both", expand=True, padx=(0, 12))
            self._card_header(card_tpl, "Word Template",
                              "The .docx file containing [placeholder] tags")

            tpl_inner = tk.Frame(card_tpl, bg=C["panel"])
            tpl_inner.pack(fill="x", padx=16, pady=(12, 6))

            self._ft_tpl_path = None
            self._ft_tpl_var  = tk.StringVar(value="No template selected")
            tk.Label(tpl_inner, textvariable=self._ft_tpl_var,
                     bg=C["panel"], fg=C["sub"],
                     font=(FF, 9), wraplength=280, justify="left"
                     ).pack(anchor="w", pady=(0, 8))
            ttk.Button(tpl_inner, text="Browse for template…",
                       style="Ghost.TButton",
                       command=self._ft_browse_template).pack(anchor="w")

            pdf_row = tk.Frame(card_tpl, bg=C["panel"])
            pdf_row.pack(fill="x", padx=16, pady=(8, 14))
            self._ft_pdf_var = tk.BooleanVar(
                value=bool(self._settings.get("ft_to_pdf", False)))
            tk.Checkbutton(pdf_row,
                           text="Also export as PDF  (requires Microsoft Word)",
                           variable=self._ft_pdf_var,
                           bg=C["panel"], fg=C["text"],
                           activebackground=C["panel"],
                           font=(FF, 9),
                           selectcolor=C["panel"]
                           ).pack(anchor="w")

            # Card B: placeholder names input
            card_ph = self._card(row1)
            card_ph.pack(side="left", fill="both", expand=True)
            self._card_header(card_ph, "Placeholders",
                              "Auto-filled from the template when you pick one — one [name] per line")

            self._ft_ph_text = scrolledtext.ScrolledText(
                card_ph, height=7, font=("Consolas", 10),
                bg="white", fg=C["text"], relief="flat",
                highlightbackground=C["border"], highlightthickness=0,
                insertbackground=C["text"], wrap="none", borderwidth=0
            )
            self._ft_ph_text.pack(fill="both", expand=True, padx=16, pady=(10, 6))

            ph_foot = tk.Frame(card_ph, bg=C["panel"])
            ph_foot.pack(fill="x", padx=16, pady=(0, 12))
            ttk.Button(ph_foot, text="Detect from template",
                       style="Ghost.TButton",
                       command=self._ft_detect_placeholders).pack(side="left")
            tk.Label(ph_foot,
                     text="no brackets, one per line",
                     bg=C["panel"], fg=C["sub"], font=(FF, 8)
                     ).pack(side="left", padx=(10, 0))

            # ── Row 2: two-step action bar ──
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(0, 14))

            self._ft_update_btn = ttk.Button(row2, text="① Update Excel",
                                             style="Primary.TButton",
                                             command=self._ft_update_excel)
            self._ft_update_btn.pack(side="left")

            tk.Label(row2, text="→  fill data in Excel, save & close  →",
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=12)

            self._ft_run_btn = ttk.Button(row2, text="② Run",
                                          style="Primary.TButton",
                                          command=self._ft_run)
            self._ft_run_btn.pack(side="left")

            self._ft_bar = ttk.Progressbar(
                row2, mode="indeterminate", length=160,
                style="Gold.Horizontal.TProgressbar"
            )

            self._ft_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._ft_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(row2, lambda: OUTPUT_FOLDER).pack(side="right")

            # ── Row 3: log ──
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._ft_log = self._log_widget(card_log)

            return frame

        def _ft_browse_template(self):
            path = filedialog.askopenfilename(
                title="Select Word template",
                filetypes=[("Word documents", "*.docx")]
            )
            if path:
                self._ft_tpl_path = path
                self._ft_tpl_var.set(f"✓  {os.path.basename(path)}")
                # Auto-detect [placeholder] tags, but never clobber names the
                # user has already typed in.
                if not self._ft_ph_text.get("1.0", "end").strip():
                    self._ft_detect_placeholders(announce=False)

        def _ft_detect_placeholders(self, announce=True):
            """Read the chosen template and fill the placeholder box with its [tags]."""
            if not self._ft_tpl_path:
                if announce:
                    messagebox.showinfo("No template",
                                        "Pick a Word template first, then detect.")
                return
            try:
                ft = _load_engine("fill_template")
                names = ft.detect_placeholders(self._ft_tpl_path)
            except Exception as e:
                if announce:
                    messagebox.showwarning(
                        "Could not read template",
                        f"Placeholders could not be detected automatically.\n\n{e}")
                return
            if not names:
                if announce:
                    messagebox.showinfo(
                        "No placeholders found",
                        "No [placeholder] tags were found in this template.\n"
                        "You can still type them in by hand.")
                return
            self._ft_ph_text.delete("1.0", "end")
            self._ft_ph_text.insert("1.0", "\n".join(names))
            self._ft_status.set(f"✓  {len(names)} placeholder(s) detected")

        def _ft_update_excel(self):
            raw   = self._ft_ph_text.get("1.0", "end")
            names = [n.strip() for n in raw.strip().splitlines() if n.strip()]

            if not names:
                messagebox.showwarning("No placeholders",
                                       "Enter at least one placeholder name.")
                return

            self._ft_update_btn.configure(state="disabled")
            self._ft_log.configure(state="normal")
            self._ft_log.delete("1.0", "end")
            self._ft_log.configure(state="disabled")

            threading.Thread(
                target=self._ft_do_update_excel, args=(names,), daemon=True
            ).start()

        def _ft_do_update_excel(self, names):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._ft_log, m))

            try:
                if SCRIPT_DIR not in sys.path:
                    sys.path.insert(0, SCRIPT_DIR)
                import fill_template as ft
                import importlib
                importlib.reload(ft)

                ok = ft.update_excel_headers(names, log=log)
                if ok:
                    self.after(0, lambda: self._ft_status.set(
                        "✓  Excel updated — fill data, then Run"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Excel updated",
                        f"{len(names)} column(s) set as headers.\n\n"
                        f"Open   code\\word_placeholder_excel.xlsx\n"
                        f"Fill in your data from row 2, save and close it,\n"
                        f"then press  ② Run."
                    ))
            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._ft_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._ft_update_btn.configure(state="normal"))

        def _ft_run(self):
            if not self._ft_tpl_path:
                messagebox.showwarning("No template", "Select a Word template first.")
                return
            if not self._check_output_folder():
                return

            to_pdf = self._ft_pdf_var.get()
            self._persist_settings()

            self._ft_run_btn.configure(state="disabled")
            self._ft_bar.pack(side="left", padx=(14, 0))
            self._ft_bar.start(8)
            self._ft_status.set("Running…")
            self._ft_log.configure(state="normal")
            self._ft_log.delete("1.0", "end")
            self._ft_log.configure(state="disabled")

            threading.Thread(
                target=self._ft_execute,
                args=(self._ft_tpl_path, to_pdf),
                daemon=True
            ).start()

        def _ft_execute(self, template_path, to_pdf):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._ft_log, m))

            com_ready = False
            try:
                if to_pdf:
                    try:
                        import pythoncom
                        pythoncom.CoInitialize()
                        com_ready = True
                    except ImportError:
                        pass

                if SCRIPT_DIR not in sys.path:
                    sys.path.insert(0, SCRIPT_DIR)
                import fill_template as ft
                import importlib
                importlib.reload(ft)

                ft.process(
                    template_path=template_path,
                    to_pdf=to_pdf,
                    log_callback=log,
                )

                self.after(0, lambda: self._ft_status.set("✓  Done"))
                self.after(0, lambda: messagebox.showinfo(
                    "Done", "All files written.\n\nOutput folder:  output/"
                ))

            except SystemExit:
                # fill_template.py calls sys.exit on data errors — already logged
                self.after(0, lambda: self._ft_status.set("Error — see log"))
            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._ft_status.set("Error — see log"))
            finally:
                if com_ready:
                    try:
                        import pythoncom
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass
                self.after(0, lambda: self._ft_run_btn.configure(state="normal"))
                self.after(0, lambda: self._ft_bar.stop())
                self.after(0, lambda: self._ft_bar.pack_forget())

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 5 — ADD PASSWORDS
        # ═══════════════════════════════════════════════════════════════════

        def _build_addpass_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            # ── Row 1: two cards side by side ──
            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="x", pady=(0, 14))

            # Card A: file / folder picker with drop zone
            card_files = self._card(row1)
            card_files.pack(side="left", fill="both", expand=True, padx=(0, 12))
            self._card_header(card_files, "Source PDFs",
                              "Select individual files or an entire folder")

            drop_wrapper = tk.Frame(card_files, bg=C["panel"], height=160)
            drop_wrapper.pack(fill="x", padx=16, pady=(10, 0))
            drop_wrapper.pack_propagate(False)

            self._ap_files = []
            self._ap_folder = None
            self._ap_files_var = tk.StringVar(value="Drop files here  ·  or use Browse below")
            ap_zone, self._ap_zone_lbl = self._dropzone(
                drop_wrapper, "📄", self._ap_files_var, self._ap_browse_files
            )
            self._register_drop(ap_zone, self._ap_on_drop)

            browse_row = tk.Frame(card_files, bg=C["panel"])
            browse_row.pack(fill="x", padx=16, pady=(8, 14))
            ttk.Button(browse_row, text="Browse files",
                       style="Ghost.TButton",
                       command=self._ap_browse_files).pack(side="left", padx=(0, 6))
            ttk.Button(browse_row, text="Browse folder",
                       style="Ghost.TButton",
                       command=self._ap_browse_folder).pack(side="left")

            # Card B: owner password
            card_pw = self._card(row1, width=280)
            card_pw.pack(side="left", fill="y")
            card_pw.pack_propagate(False)
            self._card_header(card_pw, "Owner Password",
                              "Blocks editing and copying — file still opens without a password")
            pw_inner = tk.Frame(card_pw, bg=C["panel"])
            pw_inner.pack(fill="x", padx=16, pady=(10, 14))
            self._ap_pw_var  = tk.StringVar()
            self._ap_pw_show = tk.BooleanVar(value=False)
            self._password_row(pw_inner, self._ap_pw_var, self._ap_pw_show)

            # ── Row 2: action bar ──
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(0, 14))

            self._ap_run_btn = ttk.Button(row2, text="Add Passwords",
                                          style="Primary.TButton",
                                          command=self._ap_run)
            self._ap_run_btn.pack(side="left")

            self._ap_bar = ttk.Progressbar(
                row2, mode="indeterminate", length=160,
                style="Gold.Horizontal.TProgressbar"
            )

            self._ap_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._ap_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(row2, lambda: OUTPUT_FOLDER).pack(side="right")

            # ── Row 3: log ──
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._ap_log = self._log_widget(card_log)

            return frame

        # ── Add Passwords: file handling ──────────────────────────────────

        def _ap_on_drop(self, event):
            import re
            parts = re.findall(r'\{([^}]+)\}|(\S+)', event.data.strip())
            flat  = [p[0] or p[1] for p in parts]
            pdfs  = [p for p in flat if p.lower().endswith(".pdf")]
            if pdfs:
                self._ap_set_files(pdfs)
            else:
                messagebox.showwarning("No PDFs", "No PDF files detected.")

        def _ap_browse_files(self, event=None):
            paths = filedialog.askopenfilenames(
                title="Select PDFs to protect",
                filetypes=[("PDF files", "*.pdf")]
            )
            if paths:
                self._ap_set_files(list(paths))

        def _ap_browse_folder(self):
            folder = filedialog.askdirectory(title="Select folder with PDFs")
            if not folder:
                return
            pdfs = [os.path.join(folder, f)
                    for f in os.listdir(folder) if f.lower().endswith(".pdf")]
            if not pdfs:
                messagebox.showwarning("No PDFs", "No PDF files found in that folder.")
                return
            self._ap_set_files(pdfs, folder=folder)

        def _ap_set_files(self, paths, folder=None):
            self._ap_files  = paths
            self._ap_folder = folder
            self._ap_files_var.set(f"✓  {len(paths)} file(s) selected")
            self._ap_zone_lbl.configure(fg=C["ok"])

        # ── Add Passwords: run ────────────────────────────────────────────

        def _ap_run(self):
            if not self._ap_files:
                messagebox.showwarning("No files", "Select at least one PDF.")
                return
            pw = self._ap_pw_var.get().strip()
            if not pw:
                messagebox.showwarning("Missing password", "Enter an owner password.")
                return

            ap_output = os.path.join(PARENT_DIR, "output")
            if not self._ap_check_output(ap_output):
                return

            self._ap_run_btn.configure(state="disabled")
            self._ap_bar.pack(side="left", padx=(14, 0))
            self._ap_bar.start(8)
            self._ap_status.set("Running…")
            self._ap_log.configure(state="normal")
            self._ap_log.delete("1.0", "end")
            self._ap_log.configure(state="disabled")

            threading.Thread(
                target=self._ap_execute,
                args=(list(self._ap_files), pw, ap_output),
                daemon=True
            ).start()

        def _ap_check_output(self, folder):
            """Same logic as _check_output_folder but for a specific path."""
            if not os.path.isdir(folder):
                return True
            existing = [f for f in os.listdir(folder)
                        if os.path.isfile(os.path.join(folder, f))]
            if not existing:
                return True
            ok = messagebox.askyesno(
                "Output folder is not empty",
                f"The output folder contains {len(existing)} file(s) from a previous run.\n\n"
                "Have you already moved them to a safe location?\n\n"
                "Yes  →  clear the output folder and continue\n"
                "No   →  cancel",
                icon="warning"
            )
            if not ok:
                return False
            for f in existing:
                try:
                    os.remove(os.path.join(folder, f))
                except Exception:
                    pass
            return True

        def _ap_execute(self, files, pw, output_folder):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._ap_log, m))

            try:
                if SCRIPT_DIR not in sys.path:
                    sys.path.insert(0, SCRIPT_DIR)
                import add_passwords
                import importlib
                importlib.reload(add_passwords)

                ok, total = add_passwords.protect_files(
                    files, pw, output_folder, log=log)

                if ok == total:
                    self.after(0, lambda: self._ap_status.set(f"✓  {ok} file(s) done"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Done", f"{ok} file(s) protected.\n\nOutput folder:  output/"
                    ))
                else:
                    self.after(0, lambda: self._ap_status.set(
                        f"{ok}/{total} done — check log"))

            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._ap_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._ap_run_btn.configure(state="normal"))
                self.after(0, lambda: self._ap_bar.stop())
                self.after(0, lambda: self._ap_bar.pack_forget())

        # ═══════════════════════════════════════════════════════════════════
        # SCREEN 6 — WORD TO PDF
        # ═══════════════════════════════════════════════════════════════════

        def _build_wordtopdf_screen(self, parent):
            frame = tk.Frame(parent, bg=C["bg"])

            # ── Row 1: file / folder picker card ──
            row1 = tk.Frame(frame, bg=C["bg"])
            row1.pack(fill="x", pady=(0, 14))

            card_files = self._card(row1)
            card_files.pack(side="left", fill="both", expand=True)
            self._card_header(card_files, "Files to convert",
                              "Word, Excel, PowerPoint, images, text or Outlook emails — files or a whole folder")

            drop_wrapper = tk.Frame(card_files, bg=C["panel"], height=160)
            drop_wrapper.pack(fill="x", padx=16, pady=(10, 0))
            drop_wrapper.pack_propagate(False)

            self._wp_files = []
            self._wp_folder = None   # set when a whole folder is chosen
            self._wp_last_out = None  # last folder PDFs were written to
            self._wp_files_var = tk.StringVar(
                value="Drop files here  ·  or use Browse below")
            wp_zone, self._wp_zone_lbl = self._dropzone(
                drop_wrapper, "📄", self._wp_files_var, self._wp_browse_files
            )
            self._register_drop(wp_zone, self._wp_on_drop)

            browse_row = tk.Frame(card_files, bg=C["panel"])
            browse_row.pack(fill="x", padx=16, pady=(8, 14))
            ttk.Button(browse_row, text="Browse files",
                       style="Ghost.TButton",
                       command=self._wp_browse_files).pack(side="left", padx=(0, 6))
            ttk.Button(browse_row, text="Browse folder",
                       style="Ghost.TButton",
                       command=self._wp_browse_folder).pack(side="left")

            # ── Options row ──
            opt_row = tk.Frame(frame, bg=C["bg"])
            opt_row.pack(fill="x", pady=(0, 10))
            self._wp_strip_index = tk.BooleanVar(
                value=bool(self._settings.get("wp_strip_index", True)))
            tk.Checkbutton(
                opt_row,
                text="Remove dataroom index numbers from names   "
                     "( 1.1.8.2.12 GI-GM handelshuur af 2018.pdf  →  GI-GM handelshuur af 2018.pdf )",
                variable=self._wp_strip_index,
                bg=C["bg"], fg=C["text"],
                activebackground=C["bg"], activeforeground=C["text"],
                selectcolor=C["panel"], font=(FF, 9), cursor="hand2",
            ).pack(anchor="w")
            self._wp_recurse = tk.BooleanVar(
                value=bool(self._settings.get("wp_recurse", False)))
            tk.Checkbutton(
                opt_row,
                text="Include subfolders when a folder is selected",
                variable=self._wp_recurse,
                bg=C["bg"], fg=C["text"],
                activebackground=C["bg"], activeforeground=C["text"],
                selectcolor=C["panel"], font=(FF, 9), cursor="hand2",
            ).pack(anchor="w", pady=(4, 0))
            self._wp_skip_pdfs = tk.BooleanVar(
                value=bool(self._settings.get("wp_skip_pdfs", False)))
            tk.Checkbutton(
                opt_row,
                text="Skip existing PDFs  (only the converted files go to the "
                     f"{PDF_SUBFOLDER} folder)",
                variable=self._wp_skip_pdfs,
                bg=C["bg"], fg=C["text"],
                activebackground=C["bg"], activeforeground=C["text"],
                selectcolor=C["panel"], font=(FF, 9), cursor="hand2",
            ).pack(anchor="w", pady=(4, 0))

            # ── Row 2: action bar ──
            row2 = tk.Frame(frame, bg=C["bg"])
            row2.pack(fill="x", pady=(0, 14))

            self._wp_run_btn = ttk.Button(row2, text="Convert to PDF",
                                          style="Primary.TButton",
                                          command=self._wp_run)
            self._wp_run_btn.pack(side="left")

            self._wp_bar = ttk.Progressbar(
                row2, mode="indeterminate", length=160,
                style="Gold.Horizontal.TProgressbar"
            )

            self._wp_status = tk.StringVar(value="")
            tk.Label(row2, textvariable=self._wp_status,
                     bg=C["bg"], fg=C["sub"], font=(FF, 9)
                     ).pack(side="left", padx=16)

            self._open_output_button(
                row2, lambda: self._wp_last_out).pack(side="right")

            # ── Row 3: log ──
            card_log = self._card(frame)
            card_log.pack(fill="both", expand=True)
            self._card_header(card_log, "Output Log")
            self._wp_log = self._log_widget(card_log)

            return frame

        # ── Word to PDF: file handling ────────────────────────────────────

        def _wp_on_drop(self, event):
            import re
            parts = re.findall(r'\{([^}]+)\}|(\S+)', event.data.strip())
            flat  = [p[0] or p[1] for p in parts]
            files = [p for p in flat if os.path.isfile(p)]
            if files:
                self._wp_set_files(files)
            else:
                messagebox.showwarning("No files", "No files detected.")

        def _wp_browse_files(self, event=None):
            paths = filedialog.askopenfilenames(
                title="Select files to convert",
                filetypes=[("All files", "*.*")]
            )
            if paths:
                self._wp_set_files(list(paths))

        def _wp_browse_folder(self):
            folder = filedialog.askdirectory(title="Select folder to convert")
            if not folder:
                return
            out_sub = os.path.abspath(os.path.join(folder, PDF_SUBFOLDER))
            if self._wp_recurse.get():
                files = []
                for root, dirs, names in os.walk(folder):
                    # never descend into our own pdf/ output folder
                    dirs[:] = [d for d in dirs
                               if os.path.abspath(os.path.join(root, d)) != out_sub]
                    for n in names:
                        files.append(os.path.join(root, n))
            else:
                files = [os.path.join(folder, f)
                         for f in os.listdir(folder)
                         if os.path.isfile(os.path.join(folder, f))]
            if not files:
                messagebox.showwarning("Empty folder", "No files found in that folder.")
                return
            self._wp_set_files(files, folder=folder)

        def _wp_set_files(self, paths, folder=None):
            self._wp_files  = paths
            self._wp_folder = folder
            if folder:
                self._wp_files_var.set(
                    f"✓  {len(paths)} file(s)  ·  PDFs → {os.path.basename(folder)}\\{PDF_SUBFOLDER}")
            else:
                self._wp_files_var.set(
                    f"✓  {len(paths)} file(s)  ·  PDFs saved next to each source")
            self._wp_zone_lbl.configure(fg=C["ok"])

        # ── Word to PDF: run ──────────────────────────────────────────────

        def _wp_run(self):
            if not self._wp_files:
                messagebox.showwarning("No files", "Select at least one file.")
                return

            # When a whole folder was chosen, write the PDFs into a clean
            # "_pdf" subfolder of it. For loose files, keep them next to source.
            out_dir = os.path.join(self._wp_folder, PDF_SUBFOLDER) if self._wp_folder else None
            strip_index = self._wp_strip_index.get()
            skip_pdfs   = self._wp_skip_pdfs.get()

            # Remember where output goes so "Open output folder" works after the run.
            self._wp_last_out = out_dir or (
                os.path.dirname(self._wp_files[0]) if self._wp_files else None)
            self._persist_settings()

            self._wp_run_btn.configure(state="disabled")
            self._wp_bar.pack(side="left", padx=(14, 0))
            self._wp_bar.start(8)
            self._wp_status.set("Running…")
            self._wp_log.configure(state="normal")
            self._wp_log.delete("1.0", "end")
            self._wp_log.configure(state="disabled")

            threading.Thread(
                target=self._wp_execute,
                args=(list(self._wp_files), out_dir, strip_index, skip_pdfs),
                daemon=True
            ).start()

        def _wp_execute(self, files, out_dir, strip_index, skip_pdfs):
            def log(msg):
                self.after(0, lambda m=msg: self._log_write(self._wp_log, m))

            def progress(i, total):
                pass  # indeterminate bar; per-file lines already show progress

            try:
                w2p = _load_engine("word_to_pdf")
                # out_dir set -> the _pdf subfolder; None -> next to each source file.
                if out_dir:
                    log(f"  Output folder: {out_dir}")
                if strip_index:
                    log("  Removing dataroom index numbers from output names")
                if skip_pdfs:
                    log("  Skipping existing PDFs (only converted files go to output)")
                done, skipped, failed = w2p.run_batch(
                    files, out_dir, log, progress,
                    strip_index=strip_index, skip_existing_pdfs=skip_pdfs)
                log("")
                log(f"  Finished: {done} converted, {skipped} skipped, {len(failed)} failed.")
                summary = f"{done} converted, {skipped} skipped, {len(failed)} failed."
                where = (f"in the '{PDF_SUBFOLDER}' subfolder"
                         if out_dir else "next to each source file")

                if failed:
                    self.after(0, lambda: self._wp_status.set(summary))
                    self.after(0, lambda: messagebox.showwarning(
                        "Finished with errors",
                        summary + "\n\nSee the log for details."))
                else:
                    self.after(0, lambda: self._wp_status.set(f"✓  {done} file(s) done"))
                    self.after(0, lambda w=where: messagebox.showinfo(
                        "Done",
                        f"{done} file(s) converted.\n\nEach PDF is saved {w}."))

            except Exception as e:
                log(f"  ERROR  {e}")
                self.after(0, lambda: self._wp_status.set("Error — see log"))
            finally:
                self.after(0, lambda: self._wp_run_btn.configure(state="normal"))
                self.after(0, lambda: self._wp_bar.stop())
                self.after(0, lambda: self._wp_bar.pack_forget())

    return App


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

def main():
    if sys.version_info < (3, 8):
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(
                0,
                "Python 3.8 or higher is required.\n\n"
                "Download from: https://www.python.org/downloads/\n"
                "Check 'Add Python to PATH' during installation.",
                "Python version too old", 0x10
            )
        except Exception:
            pass
        sys.exit(1)

    _pre = tk.Tk()
    _pre.withdraw()
    check_and_install_packages()
    _pre.destroy()

    try:
        from tkinterdnd2 import TkinterDnD
        AppClass = _build_app(TkinterDnD.Tk, has_dnd=True)
    except ImportError:
        AppClass = _build_app(tk.Tk, has_dnd=False)

    app = AppClass()
    app.mainloop()


if __name__ == "__main__":
    main()
